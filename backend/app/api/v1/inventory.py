from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from decimal import Decimal
import csv
import io
from openpyxl import load_workbook
from typing import List, Dict, Any

from app.core.database import get_db
from app.api.deps import get_current_user
from app.models.models import (
    User, Product, Category, Warehouse, Inventory, InventoryMovement, 
    Combo, ComboItem, Tenant
)
from app.schemas.inventory import (
    ProductCreate, ProductResponse, ProductUpdate, 
    CategoryCreate, CategoryResponse, ComboCreate, ComboResponse,
    InventoryResponse, InventoryReserve, InventoryRelease
)

router = APIRouter()

# Helper para obtener la tasa del inquilino actual
async def _get_tenant_exchange_rate(db: AsyncSession, tenant_id: Any) -> Decimal:
    result = await db.execute(select(Tenant.current_exchange_rate).where(Tenant.id == tenant_id))
    rate = result.scalar()
    return rate if rate is not None else Decimal("36.50")

# Helper para calcular el precio en VES de un producto
def _calculate_ves_price(price_usd: Decimal, price_ves_manual: Optional[Decimal], rate: Decimal) -> Decimal:
    if price_ves_manual is not None:
        return price_ves_manual
    return price_usd * rate


# --- Endpoints de Productos ---

@router.get("/products", response_model=List[ProductResponse])
async def list_products(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Lista todos los productos del Tenant (protegido por RLS).
    Calcula dinámicamente el precio en VES según la tasa del día del Tenant.
    """
    rate = await _get_tenant_exchange_rate(db, current_user.tenant_id)
    
    result = await db.execute(select(Product).where(Product.is_active == True))
    products = result.scalars().all()
    
    response = []
    for p in products:
        calculated_ves = _calculate_ves_price(p.price_usd, p.price_ves_manual, rate)
        
        # Mapeamos a ProductResponse incluyendo la columna calculada
        resp = ProductResponse(
            id=p.id,
            tenant_id=p.tenant_id,
            category_id=p.category_id,
            barcode=p.barcode,
            name=p.name,
            description=p.description,
            cost_usd=p.cost_usd,
            price_usd=p.price_usd,
            price_ves_manual=p.price_ves_manual,
            price_ves_calculated=calculated_ves,
            is_active=p.is_active,
            created_at=p.created_at,
            updated_at=p.updated_at
        )
        response.append(resp)
        
    return response


@router.post("/products", response_model=ProductResponse, status_code=status.HTTP_201_CREATED)
async def create_product(
    payload: ProductCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Crea un nuevo producto en el inventario e inicializa las existencias en el Almacén Principal.
    """
    # 1. Verificar duplicado de código de barras en el mismo tenant
    if payload.barcode:
        barcode_check = await db.execute(
            select(Product).where(Product.barcode == payload.barcode)
        )
        if barcode_check.scalars().first():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El código de barras ya existe en este comercio."
            )

    # 2. Obtener o crear Almacén Principal por defecto
    wh_query = await db.execute(
        select(Warehouse).where(Warehouse.name == "Almacén Principal")
    )
    warehouse = wh_query.scalars().first()
    if not warehouse:
        warehouse = Warehouse(tenant_id=current_user.tenant_id, name="Almacén Principal", is_active=True)
        db.add(warehouse)
        await db.flush()

    # 3. Crear Producto
    new_product = Product(
        tenant_id=current_user.tenant_id,
        category_id=payload.category_id,
        barcode=payload.barcode,
        name=payload.name,
        description=payload.description,
        cost_usd=payload.cost_usd,
        price_usd=payload.price_usd,
        price_ves_manual=payload.price_ves_manual,
        is_active=True
    )
    db.add(new_product)
    await db.flush()

    # 4. Inicializar Stock en Almacén
    new_inventory = Inventory(
        tenant_id=current_user.tenant_id,
        product_id=new_product.id,
        warehouse_id=warehouse.id,
        stock_available=payload.stock_inicial,
        stock_reserved=Decimal("0.00")
    )
    db.add(new_inventory)

    # 5. Registrar Movimiento inicial (Kardex)
    if payload.stock_inicial > 0:
        movement = InventoryMovement(
            tenant_id=current_user.tenant_id,
            product_id=new_product.id,
            warehouse_id=warehouse.id,
            quantity=payload.stock_inicial,
            type="ENTRADA",
            notes="Carga inicial de producto nuevo"
        )
        db.add(movement)

    await db.commit()

    rate = await _get_tenant_exchange_rate(db, current_user.tenant_id)
    calculated_ves = _calculate_ves_price(new_product.price_usd, new_product.price_ves_manual, rate)

    return ProductResponse(
        id=new_product.id,
        tenant_id=new_product.tenant_id,
        category_id=new_product.category_id,
        barcode=new_product.barcode,
        name=new_product.name,
        description=new_product.description,
        cost_usd=new_product.cost_usd,
        price_usd=new_product.price_usd,
        price_ves_manual=new_product.price_ves_manual,
        price_ves_calculated=calculated_ves,
        is_active=new_product.is_active,
        created_at=new_product.created_at,
        updated_at=new_product.updated_at
    )


# --- Carga Masiva (Excel / CSV) ---

@router.post("/upload", status_code=status.HTTP_201_CREATED)
async def upload_inventory(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Carga masiva de inventario a partir de un archivo Excel (.xlsx) o CSV.
    """
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""
    
    rows: List[Dict[str, Any]] = []

    try:
        if filename.endswith(".csv"):
            # Leer CSV
            text_data = content.decode("utf-8")
            csv_reader = csv.reader(io.StringIO(text_data))
            header = next(csv_reader, None) # Leer cabecera
            # Formato esperado: nombre, codigo_barras, costo_usd, precio_usd, precio_ves_manual, stock_inicial
            for row in csv_reader:
                if not row or len(row) < 1:
                    continue
                rows.append({
                    "name": row[0].strip(),
                    "barcode": row[1].strip() if len(row) > 1 and row[1].strip() else None,
                    "cost_usd": Decimal(row[2].strip() or "0.00") if len(row) > 2 else Decimal("0.00"),
                    "price_usd": Decimal(row[3].strip() or "0.00") if len(row) > 3 else Decimal("0.00"),
                    "price_ves_manual": Decimal(row[4].strip() or "0.00") if len(row) > 4 and row[4].strip() else None,
                    "stock_inicial": Decimal(row[5].strip() or "0.00") if len(row) > 5 else Decimal("0.00"),
                })
        elif filename.endswith(".xlsx"):
            # Leer Excel usando openpyxl
            wb = load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            # Recorrer filas omitiendo cabecera
            for r in list(sheet.iter_rows(values_only=True))[1:]:
                if not r or not r[0]:
                    continue
                rows.append({
                    "name": str(r[0]).strip(),
                    "barcode": str(r[1]).strip() if len(r) > 1 and r[1] is not None else None,
                    "cost_usd": Decimal(str(r[2] or "0.00")),
                    "price_usd": Decimal(str(r[3] or "0.00")),
                    "price_ves_manual": Decimal(str(r[4])) if len(r) > 4 and r[4] is not None else None,
                    "stock_inicial": Decimal(str(r[5] or "0.00")) if len(r) > 5 and r[5] is not None else Decimal("0.00"),
                })
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Formato de archivo no soportado. Debe ser .xlsx o .csv"
            )
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Error al analizar el archivo: {str(e)}"
        )

    # Buscar/crear almacén principal
    wh_query = await db.execute(select(Warehouse).where(Warehouse.name == "Almacén Principal"))
    warehouse = wh_query.scalars().first()
    if not warehouse:
        warehouse = Warehouse(tenant_id=current_user.tenant_id, name="Almacén Principal", is_active=True)
        db.add(warehouse)
        await db.flush()

    created_count = 0
    skipped_count = 0

    for r_data in rows:
        # Verificar si ya existe por código de barra en este tenant
        if r_data["barcode"]:
            dup_query = await db.execute(select(Product).where(Product.barcode == r_data["barcode"]))
            if dup_query.scalars().first():
                skipped_count += 1
                continue

        # Crear Producto
        new_prod = Product(
            tenant_id=current_user.tenant_id,
            barcode=r_data["barcode"],
            name=r_data["name"],
            cost_usd=r_data["cost_usd"],
            price_usd=r_data["price_usd"],
            price_ves_manual=r_data["price_ves_manual"],
            is_active=True
        )
        db.add(new_prod)
        await db.flush()

        # Inicializar Stock en Almacén
        new_inv = Inventory(
            tenant_id=current_user.tenant_id,
            product_id=new_prod.id,
            warehouse_id=warehouse.id,
            stock_available=r_data["stock_inicial"],
            stock_reserved=Decimal("0.00")
        )
        db.add(new_inv)

        # Kardex
        if r_data["stock_inicial"] > 0:
            mov = InventoryMovement(
                tenant_id=current_user.tenant_id,
                product_id=new_prod.id,
                warehouse_id=warehouse.id,
                quantity=r_data["stock_inicial"],
                type="ENTRADA",
                notes="Carga inicial masiva de inventario"
            )
            db.add(mov)
            
        created_count += 1

    await db.commit()
    return {"created": created_count, "skipped": skipped_count}


# --- Reserva y Liberación de Stock ---

@router.post("/reserve", response_model=InventoryResponse)
async def reserve_stock(
    payload: InventoryReserve,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Mueve cantidades de stock_available a stock_reserved.
    Lanza error si no hay suficientes existencias disponibles.
    """
    # Consultar registro de inventario
    inv_query = await db.execute(
        select(Inventory).where(
            (Inventory.product_id == payload.product_id) & 
            (Inventory.warehouse_id == payload.warehouse_id)
        )
    )
    inv = inv_query.scalars().first()
    if not inv:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventario no encontrado para este producto y almacén."
        )

    # Validar disponibilidad
    if inv.stock_available < payload.quantity:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "code": "INSUFFICIENT_STOCK",
                "message": f"No hay suficiente stock. Solicitado: {payload.quantity}, Disponible: {inv.stock_available}"
            }
        )

    # Mover stock
    inv.stock_available -= payload.quantity
    inv.stock_reserved += payload.quantity

    # Registrar movimiento
    movement = InventoryMovement(
        tenant_id=current_user.tenant_id,
        product_id=payload.product_id,
        warehouse_id=payload.warehouse_id,
        quantity=payload.quantity,
        type="RESERVA",
        reference_document=payload.sale_id,
        notes=f"Reserva para venta ID: {payload.sale_id}"
    )
    db.add(movement)
    await db.commit()
    return inv


@router.post("/release", response_model=InventoryResponse)
async def release_stock(
    payload: InventoryRelease,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Libera stock reservado devolviéndolo a disponible.
    """
    inv_query = await db.execute(
        select(Inventory).where(
            (Inventory.product_id == payload.product_id) & 
            (Inventory.warehouse_id == payload.warehouse_id)
        )
    )
    inv = inv_query.scalars().first()
    if not inv:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventario no encontrado."
        )

    # Validar que exista stock reservado para liberar
    if inv.stock_reserved < payload.quantity:
        # Si es menor, ajustamos para liberar el máximo reservado posible
        payload.quantity = inv.stock_reserved

    # Mover stock de vuelta
    inv.stock_reserved -= payload.quantity
    inv.stock_available += payload.quantity

    # Registrar movimiento
    movement = InventoryMovement(
        tenant_id=current_user.tenant_id,
        product_id=payload.product_id,
        warehouse_id=payload.warehouse_id,
        quantity=payload.quantity,
        type="LIBERACION",
        reference_document=payload.sale_id,
        notes=f"Liberación de reserva por anulación de venta ID: {payload.sale_id}"
    )
    db.add(movement)
    await db.commit()
    return inv


# --- Endpoints de Combos ---

@router.post("/combos", response_model=ComboResponse, status_code=status.HTTP_201_CREATED)
async def create_combo(
    payload: ComboCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Crea un nuevo Combo comercial con sus productos componentes.
    """
    # 1. Validar que todos los productos componentes existan y pertenezcan a este tenant
    for item in payload.items:
        prod_check = await db.execute(select(Product).where(Product.id == item.product_id))
        if not prod_check.scalars().first():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"El producto componente con ID {item.product_id} no existe."
            )

    # 2. Crear cabecera del Combo
    new_combo = Combo(
        tenant_id=current_user.tenant_id,
        name=payload.name,
        price_usd=payload.price_usd,
        is_active=True
    )
    db.add(new_combo)
    await db.flush()

    # 3. Vincular productos componentes
    for item in payload.items:
        combo_item = ComboItem(
            combo_id=new_combo.id,
            product_id=item.product_id,
            quantity=item.quantity
        )
        db.add(combo_item)
    
    await db.commit()
    
    # Consultar los items asociados de manera explícita para evitar lazy loading
    items_query = await db.execute(select(ComboItem).where(ComboItem.combo_id == new_combo.id))
    items_list = items_query.scalars().all()

    return ComboResponse(
        id=new_combo.id,
        tenant_id=new_combo.tenant_id,
        name=new_combo.name,
        price_usd=new_combo.price_usd,
        is_active=new_combo.is_active,
        created_at=new_combo.created_at,
        items=[ComboItemResponse(product_id=i.product_id, quantity=i.quantity) for i in items_list]
    )
